import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Base directory
BASE_DIR = Path(__file__).resolve().parent

# Input file
INPUT_FILE = BASE_DIR / "Crude_Petroleum_Exports_1996_2020.xlsx"

# Load merged dataset
df = pd.read_excel(INPUT_FILE)

# Ensure Trade Value is numeric
df["Trade Value"] = pd.to_numeric(df["Trade Value"], errors="coerce")

# Drop rows with missing Trade Value or Continent
df = df.dropna(subset=["Trade Value", "Continent"])

# Get all continents
continents = df["Continent"].unique()

results = []

# Analyze each continent
for continent in continents:
    subset = df[df["Continent"] == continent]["Trade Value"]

    mean_value = subset.mean()
    std_value = subset.std()
    min_value = subset.min()
    max_value = subset.max()
    range_value = max_value - min_value

    results.append({
        "Continent": continent,
        "Mean": mean_value,
        "Standard Deviation": std_value,
        "Minimum": min_value,
        "Maximum": max_value,
        "Range": range_value
    })

# Create summary table
summary_df = pd.DataFrame(results)

# Volatility classification using percentiles
low_threshold = summary_df["Range"].quantile(0.33)
high_threshold = summary_df["Range"].quantile(0.66)

def classify_volatility(value):
    if value <= low_threshold:
        return "Low volatility"
    elif value <= high_threshold:
        return "Moderate volatility"
    else:
        return "High volatility"

summary_df["Export Volatility"] = summary_df["Range"].apply(classify_volatility)

# Save to Excel
output_file = BASE_DIR / "continent_analysis_summary_v1.xlsx"
summary_df.to_excel(output_file, index=False)

# -----------------------------
# Apply color formatting
# -----------------------------

wb = load_workbook(output_file)
ws = wb.active

# Define colors
green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Find column index for Export Volatility
volatility_col = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == "Export Volatility":
        volatility_col = col
        break

# Apply colors
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=volatility_col)

    if cell.value == "Low volatility":
        cell.fill = green
    elif cell.value == "Moderate volatility":
        cell.fill = yellow
    elif cell.value == "High volatility":
        cell.fill = red

wb.save(output_file)

# Display results
print("\nCrude Petroleum Export Analysis by Continent")
print("=" * 55)
print(summary_df)

print("\nSummary saved with color formatting:", output_file)
